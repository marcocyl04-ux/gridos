"""Workbook endpoints: sheets, save/load, export/import, multi-workbook, chat log, debug, clear."""
from __future__ import annotations
import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Body, Depends, HTTPException, Query, Response, UploadFile, File
from pydantic import BaseModel
from core.engine import GridOSKernel
from core.api.deps import (
    kernel, current_kernel_dep, _scope_for, workbook_store,
    _require_saas_storage,
    cloud_config, AuthUser, require_user,
    import_file, auto_detect_template,
    WorkbookRenameRequest, WorkbookCreateRequest, WorkbookRenameSaasRequest,
    SheetCreateRequest, SheetRenameRequest, SheetActivateRequest,
    ChatLogReplaceRequest, CellClearRequest, CellFormatRequest,
    FormulaRequest,
)

router = APIRouter()


# --- Debug / workbook info ---

@router.get("/debug/grid")
async def get_grid(sheet: Optional[str] = None, k: GridOSKernel = Depends(current_kernel_dep)):
    target = sheet or kernel.active_sheet
    return {"sheet": target, "cells": kernel.export_sheet(target), "charts": kernel.list_charts(target)}


@router.get("/api/workbook")
async def get_workbook(k: GridOSKernel = Depends(current_kernel_dep)):
    return {
        "workbook_name": kernel.workbook_name,
        "active_sheet": kernel.active_sheet,
        "sheets": kernel.list_sheets(),
        "chat_log": list(kernel.chat_log),
    }


# --- Chat log ---

@router.post("/workbook/chat/replace")
async def replace_chat_log(req: ChatLogReplaceRequest, k: GridOSKernel = Depends(current_kernel_dep)):
    try:
        kernel.set_chat_log(req.entries)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"status": "Success", "count": len(kernel.chat_log)}


@router.post("/workbook/chat/clear")
async def clear_chat_log(k: GridOSKernel = Depends(current_kernel_dep)):
    kernel.clear_chat_log()
    return {"status": "Success"}


# --- Workbook name ---

@router.post("/workbook/rename")
async def rename_workbook(req: WorkbookRenameRequest, k: GridOSKernel = Depends(current_kernel_dep)):
    try:
        name = kernel.rename_workbook(req.name)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"workbook_name": name}


# --- Sheets ---

@router.post("/workbook/sheet")
async def create_sheet(req: SheetCreateRequest, k: GridOSKernel = Depends(current_kernel_dep)):
    name = kernel.create_sheet(req.name)
    return {"sheet": name, "sheets": kernel.list_sheets(), "active_sheet": kernel.active_sheet}


@router.post("/workbook/sheet/rename")
async def rename_sheet(req: SheetRenameRequest, k: GridOSKernel = Depends(current_kernel_dep)):
    name = kernel.rename_sheet(req.old_name, req.new_name)
    return {"sheet": name, "sheets": kernel.list_sheets(), "active_sheet": kernel.active_sheet}


@router.post("/workbook/sheet/activate")
async def activate_sheet(req: SheetActivateRequest, k: GridOSKernel = Depends(current_kernel_dep)):
    name = kernel.activate_sheet(req.name)
    return {"sheet": name, "sheets": kernel.list_sheets(), "active_sheet": kernel.active_sheet}


# --- System: save/load ---

@router.post("/system/save")
async def save_grid(workbook_id: Optional[str] = None, user: AuthUser = Depends(require_user),
                    k: GridOSKernel = Depends(current_kernel_dep)):
    scope = _scope_for(user, workbook_id)
    workbook_store.save(scope, kernel.export_state_dict())
    return {"status": "Success", "workbook_id": scope.workbook_id}


@router.post("/system/load")
async def load_grid(workbook_id: Optional[str] = None, user: AuthUser = Depends(require_user),
                    k: GridOSKernel = Depends(current_kernel_dep)):
    scope = _scope_for(user, workbook_id)
    state = workbook_store.load(scope)
    if state is None:
        return {"status": "Error", "message": "No save file found."}
    kernel.apply_state_dict(state)
    return {"status": "Success", "workbook_id": scope.workbook_id}


# --- Multi-workbook (SaaS) ---

@router.get("/workbooks")
async def list_workbooks(user: AuthUser = Depends(require_user)):
    _require_saas_storage()
    items = workbook_store.list(user.id)
    tier = "free"
    try:
        from cloud import usage as cloud_usage
        summary = cloud_usage.get_tier_and_usage(user.id)
        tier = summary.get("tier") or "free"
    except Exception:
        pass
    limit = cloud_config.max_workbooks(tier)
    return {
        "workbooks": items, "tier": tier, "used": len(items), "limit": limit,
        "remaining": None if limit == 0 else max(0, limit - len(items)),
    }


@router.post("/workbooks")
async def create_workbook(req: WorkbookCreateRequest, user: AuthUser = Depends(require_user)):
    _require_saas_storage()
    tier = "free"
    try:
        from cloud import usage as cloud_usage
        summary = cloud_usage.get_tier_and_usage(user.id)
        tier = summary.get("tier") or "free"
    except Exception:
        pass
    limit = cloud_config.max_workbooks(tier)
    current = workbook_store.count(user.id)
    if limit > 0 and current >= limit:
        raise HTTPException(status_code=402, detail={
            "message": f"Workbook slot cap reached ({current}/{limit}) for your tier.",
            "usage": {"tier": tier, "used": current, "limit": limit},
        })
    created = workbook_store.create_empty(user.id, req.title or "Untitled workbook")
    return {"status": "Success", **created}


@router.patch("/workbooks/{workbook_id}")
async def rename_workbook_saas(workbook_id: str, req: WorkbookRenameSaasRequest,
                                user: AuthUser = Depends(require_user)):
    _require_saas_storage()
    scope = _scope_for(user, workbook_id)
    workbook_store.rename(scope, req.title)
    return {"status": "Success", "workbook_id": workbook_id, "title": req.title.strip()[:120]}


@router.delete("/workbooks/{workbook_id}")
async def delete_workbook(workbook_id: str, user: AuthUser = Depends(require_user)):
    _require_saas_storage()
    scope = _scope_for(user, workbook_id)
    workbook_store.delete(scope)
    return {"status": "Success", "workbook_id": workbook_id}


# --- Export ---

@router.get("/system/export")
async def export_workbook(k: GridOSKernel = Depends(current_kernel_dep)):
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "-" for c in kernel.workbook_name).strip() or "workbook"
    safe_name = safe_name.replace(" ", "_")
    filename = f"{safe_name}-{timestamp}.gridos"
    body = json.dumps(kernel.export_state_dict(), indent=2)
    return Response(content=body, media_type="application/json",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/system/export.xlsx")
async def export_workbook_xlsx(k: GridOSKernel = Depends(current_kernel_dep)):
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl is not installed. Run `pip install openpyxl`.")
    state = kernel.export_state_dict()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_order = state.get("sheet_order") or list((state.get("sheets") or {}).keys()) or ["Sheet1"]
    sheets_data = state.get("sheets") or {}
    for sheet_name in sheet_order:
        safe_sheet = sheet_name[:31]
        for ch in "[]:*?/\\":
            safe_sheet = safe_sheet.replace(ch, "_")
        ws = wb.create_sheet(title=safe_sheet or "Sheet")
        cells = (sheets_data.get(sheet_name) or {}).get("cells") or {}
        for a1, cell in cells.items():
            if not isinstance(cell, dict):
                continue
            formula, value, datatype = cell.get("formula"), cell.get("value"), cell.get("datatype")
            if formula:
                ws[a1] = formula
            elif datatype == "num" and value not in (None, ""):
                try:
                    ws[a1] = float(value)
                except (TypeError, ValueError):
                    ws[a1] = value
            else:
                ws[a1] = value
    if not wb.worksheets:
        wb.create_sheet(title="Sheet1")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    safe_name = "".join(c if c.isalnum() or c in "-_ " else "-" for c in kernel.workbook_name).strip() or "workbook"
    safe_name = safe_name.replace(" ", "_")
    filename = f"{safe_name}-{timestamp}.xlsx"
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# --- Import ---

@router.post("/system/import.xlsx")
async def import_workbook_xlsx(file: UploadFile = File(...), k: GridOSKernel = Depends(current_kernel_dep)):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl is not installed.")
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Expected an .xlsx file.")
    try:
        raw = await file.read()
        wb = load_workbook(BytesIO(raw), data_only=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")
    sheets, sheet_order = {}, []
    for ws in wb.worksheets:
        name = ws.title or "Sheet"
        sheet_order.append(name)
        cells = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                raw_val = cell.value
                formula, value_str, datatype = None, "", "str"
                if isinstance(raw_val, str) and raw_val.startswith("="):
                    formula = raw_val
                elif isinstance(raw_val, (int, float)):
                    datatype, value_str = "num", str(raw_val)
                elif isinstance(raw_val, bool):
                    value_str = "TRUE" if raw_val else "FALSE"
                else:
                    value_str = str(raw_val)
                cells[cell.coordinate] = {"value": value_str, "formula": formula, "locked": False,
                                           "datatype": datatype, "agent_owner": "User"}
        sheets[name] = {"cells": cells, "charts": []}
    if not sheet_order:
        sheet_order, sheets = ["Sheet1"], {"Sheet1": {"cells": {}, "charts": []}}
    base = file.filename.rsplit(".", 1)[0]
    workbook_name = (base or "Imported workbook").strip()[:120] or "Imported workbook"
    state = {"workbook_name": workbook_name, "active_sheet": sheet_order[0],
             "sheet_order": sheet_order, "sheets": sheets, "chat_log": []}
    try:
        kernel.apply_state_dict(state)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not apply imported state: {e}")
    return {"status": "Success", "workbook_name": workbook_name, "sheets": len(sheet_order)}


@router.post("/system/import")
async def import_workbook(payload: dict = Body(...), k: GridOSKernel = Depends(current_kernel_dep)):
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Invalid workbook payload.")
    try:
        kernel.apply_state_dict(payload)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not import workbook: {e}")
    return {"status": "Success"}


@router.post("/import/file")
async def import_file_endpoint(
    file: UploadFile = File(...),
    has_header: bool = Query(default=True),
    sheet_name: Optional[str] = Query(None),
    target_sheet: str = Query("Sheet1"),
    clear_existing: bool = Query(True),
    k: GridOSKernel = Depends(current_kernel_dep),
):
    import tempfile as _tf
    contents = await file.read()
    try:
        suffix = "." + file.filename.rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else ".csv"
        with _tf.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(contents)
            tmp_path = tmp.name
        result = import_file(tmp_path)
        Path(tmp_path).unlink(missing_ok=True)
        if result.errors:
            raise HTTPException(status_code=400, detail=f"Import failed: {'; '.join(result.errors)}")
        sheets_data = []
        for s in result.sheets:
            sheet_cells = {}
            for a1_ref, icell in s.cells.items():
                sheet_cells[a1_ref] = {"value": icell.value, "formula": icell.formula,
                                        "locked": False, "datatype": icell.datatype, "agent_owner": "User"}
            sheets_data.append({"name": s.name, "cells": sheet_cells, "rows": s.rows,
                                "cols": s.cols, "data_rows": s.data_rows})
        return {
            "status": "Success", "sheets_imported": len(result.sheets),
            "detected_type": result.detected_type,
            "template_suggestion": auto_detect_template(result),
            "warnings": result.warnings, "sheets": sheets_data,
            "target_sheet": target_sheet, "clear_existing": clear_existing,
            "has_header": has_header, "sheet_name": sheet_name,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {e}")


# --- System: clear, unlock ---

@router.post("/system/clear")
async def clear_grid(sheet: Optional[str] = None, k: GridOSKernel = Depends(current_kernel_dep)):
    kernel.clear_unlocked(sheet)
    return {"status": "Success", "sheet": sheet or kernel.active_sheet}


@router.post("/system/unlock-all")
async def unlock_all(k: GridOSKernel = Depends(current_kernel_dep)):
    dropped, unlocked = 0, 0
    for sheet_name, state in kernel.sheets.items():
        cells = state["cells"]
        for coords in list(cells.keys()):
            cell = cells[coords]
            if not cell.locked:
                continue
            if cell.value in (None, "") and not cell.formula:
                del cells[coords]
                dropped += 1
            else:
                cell.locked = False
                unlocked += 1
    return {"status": "Success", "unlocked": unlocked, "dropped": dropped}
