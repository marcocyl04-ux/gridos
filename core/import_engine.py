"""
Data Import Pipeline — CSV/Excel upload → auto-parse → populate grid.

Supports:
- CSV files
- Excel (.xlsx) files
- Auto-detection of data types (numbers, formulas, text)
- Automatic header detection
- Sheet selection for multi-sheet Excel files
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None


@dataclass
class ImportedCell:
    """A single cell from an imported file."""
    value: Any
    formula: Optional[str] = None
    datatype: str = "string"  # string, int, float, boolean, formula
    row: int = 0
    col: int = 0
    a1_ref: str = ""


@dataclass
class ImportedSheet:
    """A sheet from an imported file."""
    name: str
    cells: Dict[str, ImportedCell] = field(default_factory=dict)
    rows: int = 0
    cols: int = 0
    header_row: List[str] = field(default_factory=list)
    data_rows: int = 0


@dataclass
class ImportResult:
    """Result of a file import."""
    sheets: List[ImportedSheet] = field(default_factory=list)
    detected_type: str = ""  # "financial", "tabular", "unknown"
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def import_csv(file_path: str | Path, has_header: bool = True) -> ImportResult:
    """Import a CSV file into a grid-friendly format."""
    result = ImportResult()
    result.detected_type = "tabular"
    
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
    except UnicodeDecodeError:
        # Try latin-1 encoding as fallback
        with open(file_path, "r", encoding="latin-1") as f:
            reader = csv.reader(f)
            rows = list(reader)
    
    if not rows:
        result.errors.append("CSV file is empty")
        return result
    
    sheet = ImportedSheet(name="Sheet1")
    sheet.rows = len(rows)
    sheet.cols = max(len(row) for row in rows) if rows else 0
    
    # Detect header
    if has_header and len(rows) > 1:
        sheet.header_row = [str(cell) for cell in rows[0]]
        start_row = 1
    else:
        sheet.header_row = []
        start_row = 0
    
    # Populate cells
    for r, row in enumerate(rows[start_row:], start=1):
        for c, raw_value in enumerate(row):
            if raw_value is None or raw_value.strip() == "":
                continue
            
            cell = _parse_cell_value(raw_value.strip(), r, c + 1)
            if cell:
                sheet.cells[cell.a1_ref] = cell
    
    sheet.data_rows = len(rows) - start_row
    result.sheets.append(sheet)
    
    return result


def import_excel(file_path: str | Path) -> ImportResult:
    """Import an Excel file, including formulas."""
    result = ImportResult()
    
    if not OPENPYXL_AVAILABLE:
        result.errors.append(
            "Excel import not available. Install openpyxl: pip install openpyxl"
        )
        return result
    
    try:
        wb = load_workbook(file_path, data_only=False)
    except Exception as e:
        result.errors.append(f"Could not read Excel file: {e}")
        return result
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet = ImportedSheet(name=ws_name)
        
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        
        # Read header row (row 1)
        sheet.header_row = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=1, column=col_idx).value
            sheet.header_row.append(str(val) if val is not None else "")
        
        # Read all cells
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                raw_value = cell.value
                formula = None
                
                # Check if cell has a formula
                if cell.data_type == "f" and raw_value:
                    formula = f"={raw_value}"
                    raw_value = formula  # Store formula string for re-evaluation
                
                if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
                    continue
                
                imported_cell = _create_imported_cell(
                    raw_value, 
                    formula,
                    row_idx, 
                    col_idx
                )
                if imported_cell:
                    sheet.cells[imported_cell.a1_ref] = imported_cell
        
        sheet.rows = max_row
        sheet.cols = max_col
        sheet.data_rows = max(max_row - 1, 0)
        
        if sheet.cells:  # Only add sheets that have data
            result.sheets.append(sheet)
    
    return result


def import_file(file_path: str | Path) -> ImportResult:
    """Import any supported file type."""
    path = Path(file_path)
    suffix = path.suffix.lower()
    
    if suffix == ".csv":
        return import_csv(path)
    elif suffix in (".xlsx", ".xlsm"):
        return import_excel(path)
    else:
        result = ImportResult()
        result.errors.append(f"Unsupported file type: {suffix}")
        return result


def auto_detect_template(result: ImportResult) -> Optional[str]:
    """Try to detect what type of data this is based on headers."""
    if not result.sheets:
        return None
    
    sheet = result.sheets[0]
    if not sheet.header_row:
        return None
    
    headers = [h.lower().strip() for h in sheet.header_row]
    header_text = " ".join(headers)
    
    # Financial model patterns
    financial_indicators = [
        "revenue", "cogs", "gross profit", "ebitda", "net income",
        "assets", "liabilities", "equity", "cash flow",
        "income statement", "balance sheet", "cash flow statement"
    ]
    
    if any(ind in header_text for ind in financial_indicators):
        result.detected_type = "financial"
        return "financial"
    
    # Real estate patterns
    re_indicators = [
        "noi", "cap rate", "dscr", "rent roll", "vacancy",
        "property", "rental", "mortgage"
    ]
    
    if any(ind in header_text for ind in re_indicators):
        result.detected_type = "real_estate"
        return "real_estate"
    
    return "tabular"


def _normalize_cell_ref(row: int, col: int) -> str:
    """Convert (row, col) to A1 notation. Row is 1-indexed."""
    col_letters = ""
    while col > 0:
        col -= 1
        col_letters = chr(ord("A") + (col % 26)) + col_letters
        col //= 26
    return f"{col_letters}{row}"


def _parse_cell_value(raw: str, row: int, col: int) -> Optional[ImportedCell]:
    """Parse a raw string value into a typed ImportedCell."""
    if raw is None or raw.strip() == "":
        return None
    
    a1_ref = _normalize_cell_ref(row, col)
    
    # Formula
    if raw.startswith("="):
        return ImportedCell(
            value=raw,
            formula=raw,
            datatype="formula",
            row=row,
            col=col,
            a1_ref=a1_ref
        )
    
    # Boolean
    if raw.lower() in ("true", "false"):
        return ImportedCell(
            value=raw.lower() == "true",
            datatype="boolean",
            row=row,
            col=col,
            a1_ref=a1_ref
        )
    
    # Number
    cleaned = raw.replace(",", "").replace("$", "").replace("%", "").strip()
    try:
        if "." in cleaned:
            return ImportedCell(
                value=float(cleaned),
                datatype="float",
                row=row,
                col=col,
                a1_ref=a1_ref
            )
        return ImportedCell(
            value=int(cleaned),
            datatype="int",
            row=row,
            col=col,
            a1_ref=a1_ref
        )
    except ValueError:
        pass
    
    # String
    return ImportedCell(
        value=raw,
        datatype="string",
        row=row,
        col=col,
        a1_ref=a1_ref
    )


def _create_imported_cell(raw_value: Any, formula: Optional[str], row: int, col: int) -> Optional[ImportedCell]:
    """Create an ImportedCell from a raw Excel cell value."""
    a1_ref = _normalize_cell_ref(row, col)
    
    if formula:
        return ImportedCell(
            value=formula,
            formula=formula,
            datatype="formula",
            row=row,
            col=col,
            a1_ref=a1_ref
        )
    
    if raw_value is None:
        return None
    
    if isinstance(raw_value, bool):
        return ImportedCell(
            value=raw_value,
            datatype="boolean",
            row=row,
            col=col,
            a1_ref=a1_ref
        )
    
    if isinstance(raw_value, (int, float)):
        return ImportedCell(
            value=raw_value,
            datatype="float" if isinstance(raw_value, float) else "int",
            row=row,
            col=col,
            a1_ref=a1_ref
        )
    
    return ImportedCell(
        value=str(raw_value),
        datatype="string",
        row=row,
        col=col,
        a1_ref=a1_ref
    )
